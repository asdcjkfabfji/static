import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def generate_excel(df, filename, selected_year, report_date=None):
    if report_date is None:
        report_date = datetime.today().strftime('%d.%m.%Y')
    
    df_f = df[df['year'] == selected_year].copy()
    
    # --- TA'LIM SHAKLLARINI BIRLASHTIRISH QISMI (YANGI) ---
    def normalize_edu(name):
        name = str(name).lower()
        if 'masofaviy' in name: return "Масофавий"
        if 'sirtqi' in name: return "Сиртқи"
        if 'kunduzgi' in name: return "Кундузги"
        if 'kech' in name: return "Кечки"
        return name.capitalize()

    df_f['education_form'] = df_f['education_form'].apply(normalize_edu)
    # -----------------------------------------------------

    if df_f.empty:
        raise ValueError(f"{selected_year} yil uchun ma'lumot topilmadi.")
    
    all_courses = sorted(df_f['level'].unique())
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 1. TA'LIM SHAKLLARI BO'YICHA SHEETLAR
        for idx, shakl in enumerate(df_f['education_form'].unique()):
            shakl_df = df_f[df_f['education_form'] == shakl]
            sheet_data = []
            total_cols = len(all_courses) * 7
            
            # --- SHEET UCHUN STATISTIKA YIG'ISH ---
            sheet_course_stats = {c: {"jami": 0, "qiz": 0, "ogil": 0, "guruhlar": 0} for c in all_courses}
            
            sheet_data.append([
                "Буҳоро давлат техника университетининг факультетлар кесимида талабалар контингенти"
            ] + [""] * (total_cols - 1) + [report_date])
            sheet_data.append([""] * (total_cols + 1))
            sheet_data.append([""] * (total_cols + 1)) 
            
            headers = []
            for _ in all_courses:
                headers.extend(["т/р", "Гуруҳ номи", "Жами", "Грант", "Контракт", "Қиз", "Ўғил"])
            sheet_data.append(headers + [""])
            
            for fak in shakl_df['department'].unique():
                sheet_data.append([fak] + [""] * total_cols)
                fak_df = shakl_df[shakl_df['department'] == fak]
                kurslar_dict = {}
                max_rows = 0
                course_totals = {c: {"grant":0, "kontr":0, "qiz":0, "ogil":0, "jami":0} for c in all_courses}
                
                for kurs in all_courses:
                    k_df = fak_df[fak_df['level'] == kurs]
                    if k_df.empty:
                        kurslar_dict[kurs] = []
                        continue
                    
                    # .astype(str) TypeError: 'int' is not iterable xatosini oldini olish uchun
                    stats = k_df.groupby('group_name').agg(
                        jami=('student_id', 'count'),
                        grant=('student_id', lambda x: k_df.loc[x.index, 'payment_form'].astype(str).str.contains('grant', case=False, na=False).sum()),
                        kontr=('student_id', lambda x: (~k_df.loc[x.index, 'payment_form'].astype(str).str.contains('grant', case=False, na=False)).sum()),
                        qiz=('gender', lambda x: (x == 'Ayol').sum()),
                        ogil=('gender', lambda x: (x == 'Erkak').sum())
                    ).reset_index()
                    
                    records = stats.to_dict('records')
                    kurslar_dict[kurs] = records
                    max_rows = max(max_rows, len(records))
                    
                    course_totals[kurs]["jami"] += stats['jami'].sum()
                    course_totals[kurs]["grant"] += stats['grant'].sum()
                    course_totals[kurs]["kontr"] += stats['kontr'].sum()
                    course_totals[kurs]["qiz"] += stats['qiz'].sum()
                    course_totals[kurs]["ogil"] += stats['ogil'].sum()

                    # Sheet yakuni uchun ma'lumotlarni qo'shish
                    sheet_course_stats[kurs]["jami"] += stats['jami'].sum()
                    sheet_course_stats[kurs]["qiz"] += stats['qiz'].sum()
                    sheet_course_stats[kurs]["ogil"] += stats['ogil'].sum()
                    sheet_course_stats[kurs]["guruhlar"] += len(records)
                
                for i in range(max_rows):
                    row = []
                    for kurs in all_courses:
                        groups = kurslar_dict.get(kurs, [])
                        if i < len(groups):
                            g = groups[i]
                            row.extend([i+1, g['group_name'], g['jami'], g['grant'], g['kontr'], g['qiz'], g['ogil']])
                        else:
                            row.extend([""] * 7)
                    sheet_data.append(row + [""])
                
                jami_row = ["Жами"]
                for kurs in all_courses:
                    t = course_totals[kurs]
                    jami_row.extend(["", t["jami"], t["grant"], t["kontr"], t["qiz"], t["ogil"]])
                sheet_data.append(jami_row + [""])
                sheet_data.append([""] * (total_cols + 1))
            
            # --- SHEET YAKUNIDA UMUMIY HISOBOТ QISMINI QO'SHISH ---
            sheet_data.append([""] * (total_cols + 1))
            sheet_data.append([f"{shakl} таълим шакли бўйича умумий якуний ҳисобот"] + [""] * (total_cols))
            sheet_data.append(["Курслар", "Гуруҳлар сони", "Жами талаба", "Қизлар", "Ўғиллар"] + [""] * (total_cols - 4))
            
            grand_guruhlar, grand_jami, grand_qiz, grand_ogil = 0, 0, 0, 0
            for kurs in all_courses:
                s = sheet_course_stats[kurs]
                sheet_data.append([f"{kurs}-курс", s["guruhlar"], s["jami"], s["qiz"], s["ogil"]] + [""] * (total_cols - 4))
                grand_guruhlar += s["guruhlar"]
                grand_jami += s["jami"]
                grand_qiz += s["qiz"]
                grand_ogil += s["ogil"]
            
            sheet_data.append(["Жами:", grand_guruhlar, grand_jami, grand_qiz, grand_ogil] + [""] * (total_cols - 4))
            # -----------------------------------------------------

            final_df = pd.DataFrame(sheet_data)
            sheet_name = str(shakl).replace('/', '_')[:31]
            final_df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            # FORMATLASH
            ws = writer.sheets[sheet_name]
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            light_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            dark_green = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
            gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            tnr_font = Font(name="Times New Roman", size=11)
            tnr_bold = Font(name="Times New Roman", bold=True, size=11)
            tnr_big_bold = Font(name="Times New Roman", bold=True, size=14)
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=total_cols + 1):
                cell_val = row[0].value
                row_idx = row[0].row
                
                for cell in row:
                    if cell.column <= total_cols:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.font = tnr_font
                
                if row_idx == 1:
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
                    ws.cell(1, 1).font = tnr_big_bold
                elif row_idx == 3:
                    for i, kurs in enumerate(all_courses):
                        ws.merge_cells(start_row=3, start_column=i*7+1, end_row=3, end_column=i*7+7)
                        ws.cell(3, i*7+1).value = f"{kurs}-курс"
                        ws.cell(3, i*7+1).fill = light_green
                        ws.cell(3, i*7+1).font = tnr_bold
                elif row_idx == 4:
                    for cell in row[:-1]:
                        if cell.value: cell.fill = gray_fill; cell.font = tnr_bold
                elif cell_val and len(str(cell_val)) > 8 and "курс" not in str(cell_val) and "якуний ҳисобот" not in str(cell_val):
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols)
                    for c in range(1, total_cols + 1):
                        ws.cell(row_idx, c).fill = dark_green
                        ws.cell(row_idx, c).font = Font(name="Times New Roman", bold=True, color="FFFFFF", size=12)
                elif cell_val == "Жами":
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    for cell in row[:-1]:
                        if cell.value != "": cell.fill = yellow_fill; cell.font = tnr_bold

                # YAKUNIY HISOBOТ FORMATI
                if cell_val and ("якуний ҳисобот" in str(cell_val) or cell_val == "Курслар" or cell_val == "Жами:"):
                    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=total_cols) if "якуний ҳисобот" in str(cell_val) else None
                    for cell in row[:5]:
                        cell.fill = blue_fill
                        cell.font = tnr_bold

            widths = [6, 17, 9, 9, 11, 8, 8]
            for i in range(1, total_cols + 1):
                ws.column_dimensions[get_column_letter(i)].width = widths[(i-1) % 7]

        # 2. "ЖАМИ КОНТИНГЕНТ" SHEETINI QO'SHISH (Asl holicha qoldi)
        summary_data = []
        summary_data.append([f"Буҳоро давлат техника университети талабалар контингенти {selected_year} ўқув йили"] + [""] * (len(all_courses) + 2))
        summary_data.append(["Таълим шакли"] + [f"{c}-курс" for c in all_courses] + ["Жами", "Қизлар"])
        
        grand_totals = {c: 0 for c in all_courses}
        total_all, total_girls = 0, 0
        for shakl in df_f['education_form'].unique():
            sh_df = df_f[df_f['education_form'] == shakl]
            row = [shakl]
            for c in all_courses:
                count = len(sh_df[sh_df['level'] == c])
                row.append(count if count > 0 else 0)
                grand_totals[c] += count
            row.append(len(sh_df))
            row.append(len(sh_df[sh_df['gender'] == 'Ayol']))
            summary_data.append(row)
            total_all += len(sh_df)
            total_girls += len(sh_df[sh_df['gender'] == 'Ayol'])
            
        summary_data.append(["Жами"] + [grand_totals[c] for c in all_courses] + [total_all, total_girls])
        sum_df = pd.DataFrame(summary_data)
        sum_sheet_name = "Жами контингент"
        sum_df.to_excel(writer, sheet_name=sum_sheet_name, index=False, header=False)
        
        ws_sum = writer.sheets[sum_sheet_name]
        ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(all_courses)+3)
        for row in ws_sum.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name="Times New Roman", size=12)
                if cell.row == 1: cell.font = Font(name="Times New Roman", bold=True, size=14)
                if cell.row == 2 or cell.column == 1: cell.fill = light_green; cell.font = Font(name="Times New Roman", bold=True)
                if cell.row == ws_sum.max_row: cell.fill = yellow_fill; cell.font = Font(name="Times New Roman", bold=True)
        
        ws_sum.column_dimensions['A'].width = 30
        for i in range(2, len(all_courses) + 4):
            ws_sum.column_dimensions[get_column_letter(i)].width = 15

    print(f"Excel fayl '{filename}' muvaffaqiyatli yaratildi!")